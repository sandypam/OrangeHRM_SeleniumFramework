import os
import openpyxl

# Data Driven test from an excel sheet! this tells below test to use all info in excel sheet
def get_data(sheetName):
    base_dir = os.path.dirname(__file__)
    excel_path = os.path.join(base_dir, "..", "excel", "testdata.xlsx")

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheetName]

    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2, totalrows + 1):  # skip header
        if totalcols == 1:
            # Single-column sheet → return one value per row
            value = sheet.cell(row=i, column=1).value
            if value is not None:
                mainList.append(str(value).strip())
        else:
            # Multi-column sheet → return tuple per row
            row = []
            for j in range(1, totalcols + 1):
                row.append(sheet.cell(row=i, column=j).value)
            mainList.append(tuple(row))

    return mainList
